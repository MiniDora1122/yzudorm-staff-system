from __future__ import annotations

import calendar
import csv
from collections import defaultdict
from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import joinedload

from ..extensions import db
from ..models import (
    DocumentPageKind,
    DocumentStatus,
    LeaveRequest,
    Shift,
    ShiftPublicationStatus,
    ShiftStatus,
    ShiftType,
    StaffDocument,
    StaffProfile,
    SwapRequest,
)
from .documents import expiry_state
from .payroll import calculate_staff_cost, get_payroll_setting


WEEKDAYS = ("日", "一", "二", "三", "四", "五", "六")
THIN_GRAY = Side(style="thin", color="808080")
MEDIUM_DARK = Side(style="medium", color="303030")
CALENDAR_FILL = PatternFill("solid", fgColor="D9EFD2")
HEADER_FILL = PatternFill("solid", fgColor="EAF1F8")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FILL = PatternFill("solid", fgColor="17365D")
MATRIX_HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
MATRIX_WEEKEND_FILL = PatternFill("solid", fgColor="FFF2E5")
MATRIX_HOURS_FILL = PatternFill("solid", fgColor="DDEBF7")
MATRIX_ALT_FILL = PatternFill("solid", fgColor="F7F9FC")


def _roc_year(year: int) -> int:
    return year - 1911


def _active_profiles() -> list[StaffProfile]:
    return db.session.scalars(
        db.select(StaffProfile)
        .where(StaffProfile.user.has(is_active=True))
        .order_by(StaffProfile.student_number, StaffProfile.name)
    ).all()


def _profiles_for_month(start: date, end: date) -> list[StaffProfile]:
    """Include active students and archived students who still have reportable shifts."""
    scheduled_staff = db.select(Shift.staff_id).where(
        Shift.shift_date >= start,
        Shift.shift_date < end,
        Shift.status == ShiftStatus.SCHEDULED,
        Shift.publication_status == ShiftPublicationStatus.PUBLISHED,
    )
    return db.session.scalars(
        db.select(StaffProfile)
        .where(
            db.or_(
                StaffProfile.user.has(is_active=True),
                StaffProfile.id.in_(scheduled_staff),
            )
        )
        .order_by(StaffProfile.student_number, StaffProfile.name)
    ).all()


def _month_shifts(start: date, end: date) -> list[Shift]:
    return db.session.scalars(
        db.select(Shift)
        .options(
            joinedload(Shift.staff),
            joinedload(Shift.shift_type).joinedload(ShiftType.work_location),
        )
        .join(ShiftType)
        .join(StaffProfile, Shift.staff_id == StaffProfile.id)
        .where(
            Shift.shift_date >= start,
            Shift.shift_date < end,
            Shift.status == ShiftStatus.SCHEDULED,
            Shift.publication_status == ShiftPublicationStatus.PUBLISHED,
        )
        .order_by(Shift.shift_date, ShiftType.start_time, StaffProfile.name)
    ).all()


def build_monthly_hours_workbook(start: date, end: date) -> bytes:
    """Create the reference-style per-student monthly appointment-hours calendar."""
    profiles = _profiles_for_month(start, end)
    shifts = _month_shifts(start, end)
    hours_by_staff_date: dict[tuple[int, date], Decimal] = defaultdict(Decimal)
    for shift in shifts:
        hours_by_staff_date[(shift.staff_id, shift.shift_date)] += Decimal(
            str(shift.shift_type.default_hours)
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{_roc_year(start.year)}年{start.month}月約用時數"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "C2"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35
    sheet.oddFooter.center.text = f"民國 {_roc_year(start.year)} 年 {start.month} 月｜第 &P 頁／共 &N 頁"

    widths = {"A": 12, "B": 18, "J": 12}
    for column in "CDEFGHI":
        widths[column] = 12
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    weeks = calendar.Calendar(firstweekday=calendar.SUNDAY).monthdatescalendar(
        start.year, start.month
    )
    current_row = 1
    weekly_total_cells: list[str]
    for index, profile in enumerate(profiles, start=1):
        header_row = current_row
        final_row = header_row + len(weeks) * 2 + 1
        last_calendar_row = final_row - 1
        sheet.merge_cells(start_row=header_row, start_column=1, end_row=last_calendar_row, end_column=1)
        sheet.merge_cells(start_row=header_row + 1, start_column=2, end_row=last_calendar_row, end_column=2)
        sheet.cell(header_row, 1, f"人員 {index}")
        sheet.cell(header_row, 2, f"{start.month}月")
        sheet.cell(header_row + 1, 2, f"{profile.student_number}\n{profile.name}")
        for offset, weekday in enumerate(WEEKDAYS, start=3):
            sheet.cell(header_row, offset, weekday)

        block = sheet.cell(header_row, 1)
        block.font = Font(name="Microsoft JhengHei", size=12, bold=True)
        block.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        student_cell = sheet.cell(header_row + 1, 2)
        student_cell.font = Font(name="Microsoft JhengHei", size=12)
        student_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.cell(header_row, 2).alignment = Alignment(horizontal="center", vertical="center")
        for column in range(2, 10):
            cell = sheet.cell(header_row, column)
            cell.font = Font(name="Microsoft JhengHei", size=12, bold=(column >= 3))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = HEADER_FILL

        weekly_total_cells = []
        for week_index, week in enumerate(weeks):
            date_row = header_row + 1 + week_index * 2
            hours_row = date_row + 1
            for day_index, day in enumerate(week, start=3):
                date_cell = sheet.cell(date_row, day_index)
                hours_cell = sheet.cell(hours_row, day_index)
                if day.month == start.month:
                    date_cell.value = day.day
                    hours = hours_by_staff_date.get((profile.id, day), Decimal("0"))
                    if hours:
                        hours_cell.value = int(hours) if hours == hours.to_integral_value() else float(hours)
                date_cell.fill = CALENDAR_FILL if day.month == start.month else PatternFill("solid", fgColor="F2F2F2")
                date_cell.alignment = Alignment(horizontal="right", vertical="center")
                hours_cell.alignment = Alignment(horizontal="center", vertical="center")
                hours_cell.font = Font(name="Microsoft JhengHei", size=12, bold=True, color="FF0000")
            total_cell = sheet.cell(hours_row, 10)
            total_cell.value = f"=SUM(C{hours_row}:I{hours_row})"
            total_cell.number_format = "General"
            total_cell.alignment = Alignment(horizontal="center", vertical="center")
            weekly_total_cells.append(total_cell.coordinate)
            sheet.row_dimensions[date_row].height = 22
            sheet.row_dimensions[hours_row].height = 24

        month_total = sheet.cell(final_row, 10)
        month_total.value = f"=SUM({','.join(weekly_total_cells)})"
        month_total.number_format = "General"
        month_total.font = Font(name="Microsoft JhengHei", size=12, bold=True)
        month_total.fill = TOTAL_FILL
        month_total.alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(final_row, 9, "月總計")
        sheet.cell(final_row, 9).font = Font(name="Microsoft JhengHei", size=11, bold=True)
        sheet.cell(final_row, 9).alignment = Alignment(horizontal="right", vertical="center")
        sheet.cell(final_row, 9).fill = TOTAL_FILL

        for row in range(header_row, last_calendar_row + 1):
            for column in range(1, 10):
                cell = sheet.cell(row, column)
                cell.border = Border(
                    left=MEDIUM_DARK if column == 1 else THIN_GRAY,
                    right=MEDIUM_DARK if column == 9 else THIN_GRAY,
                    top=MEDIUM_DARK if row == header_row else THIN_GRAY,
                    bottom=MEDIUM_DARK if row == last_calendar_row else THIN_GRAY,
                )
        sheet.row_dimensions[header_row].height = 24
        sheet.row_dimensions[final_row].height = 24
        current_row = final_row + 2

    if not profiles:
        sheet["A1"] = "此月份沒有可匯出的工讀生資料。"
    else:
        sheet.print_area = f"A1:J{current_row - 2}"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_daily_hours_matrix_workbook(start: date, end: date) -> bytes:
    """Create a compact all-student matrix with one column per day of the month."""
    profiles = _profiles_for_month(start, end)
    shifts = _month_shifts(start, end)
    hours_by_staff_date: dict[tuple[int, date], Decimal] = defaultdict(Decimal)
    for shift in shifts:
        hours_by_staff_date[(shift.staff_id, shift.shift_date)] += Decimal(
            str(shift.shift_type.default_hours)
        )

    days = [start.replace(day=day) for day in range(1, (end - start).days + 1)]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{_roc_year(start.year)}年{start.month}月每日時數"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "D3"

    last_day_column = 3 + len(days)
    total_column = last_day_column + 1
    last_column_letter = get_column_letter(total_column)
    title = f"工讀生約用時數月報－民國 {_roc_year(start.year)} 年 {start.month} 月"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_column)
    title_cell = sheet.cell(1, 1, title)
    title_cell.font = Font(name="Microsoft JhengHei", size=16, bold=True, color="FFFFFF")
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 32

    headers = ["序號\nNo.", "學號\nStudent ID", "姓名\nName"]
    headers.extend(f"{day.day}\n{WEEKDAYS[(day.weekday() + 1) % 7]}" for day in days)
    headers.append("小計\nTotal")
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(2, column, value)
        cell.font = Font(name="Microsoft JhengHei", size=10, bold=True, color="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = (
            MATRIX_WEEKEND_FILL
            if 4 <= column <= last_day_column and days[column - 4].weekday() >= 5
            else TOTAL_FILL if column == total_column else MATRIX_HEADER_FILL
        )
    sheet.row_dimensions[2].height = 34

    first_data_row = 3
    for index, profile in enumerate(profiles, start=1):
        row = first_data_row + index - 1
        sheet.cell(row, 1, index)
        sheet.cell(row, 2, profile.student_number)
        sheet.cell(row, 3, profile.name)
        sheet.cell(row, 2).number_format = "@"
        for day_index, day in enumerate(days, start=4):
            hours = hours_by_staff_date.get((profile.id, day), Decimal("0"))
            cell = sheet.cell(row, day_index)
            if hours:
                cell.value = int(hours) if hours == hours.to_integral_value() else float(hours)
                cell.fill = MATRIX_HOURS_FILL
                cell.font = Font(name="Microsoft JhengHei", size=10, bold=True, color="1F4E78")
            elif day.weekday() >= 5:
                cell.fill = MATRIX_WEEKEND_FILL
            elif index % 2 == 0:
                cell.fill = MATRIX_ALT_FILL
            cell.number_format = "General"
            cell.alignment = Alignment(horizontal="center", vertical="center")
        first_day_letter = get_column_letter(4)
        last_day_letter = get_column_letter(last_day_column)
        total_cell = sheet.cell(row, total_column, f"=SUM({first_day_letter}{row}:{last_day_letter}{row})")
        total_cell.number_format = "General"
        total_cell.font = Font(name="Microsoft JhengHei", size=10, bold=True, color="7F6000")
        total_cell.fill = TOTAL_FILL
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in (1, 2):
            sheet.cell(row, column).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center")
        for column in range(1, 4):
            sheet.cell(row, column).font = Font(name="Microsoft JhengHei", size=10)
            if index % 2 == 0:
                sheet.cell(row, column).fill = MATRIX_ALT_FILL
        sheet.row_dimensions[row].height = 25

    total_row = first_data_row + len(profiles)
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    sheet.cell(total_row, 1, "每日合計 Daily total")
    sheet.cell(total_row, 1).alignment = Alignment(horizontal="right", vertical="center")
    for column in range(4, total_column + 1):
        letter = get_column_letter(column)
        if profiles:
            sheet.cell(total_row, column, f"=SUM({letter}{first_data_row}:{letter}{total_row - 1})")
        else:
            sheet.cell(total_row, column, 0)
        sheet.cell(total_row, column).number_format = "General"
        sheet.cell(total_row, column).alignment = Alignment(horizontal="center", vertical="center")
    for column in range(1, total_column + 1):
        sheet.cell(total_row, column).fill = TITLE_FILL
        sheet.cell(total_row, column).font = Font(name="Microsoft JhengHei", size=10, bold=True, color="FFFFFF")
    sheet.row_dimensions[total_row].height = 27

    report_range = sheet.iter_rows(min_row=2, max_row=total_row, min_col=1, max_col=total_column)
    light_side = Side(style="thin", color="B7C9DD")
    for row in report_range:
        for cell in row:
            cell.border = Border(left=light_side, right=light_side, top=light_side, bottom=light_side)

    sheet.column_dimensions["A"].width = 7
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 18
    for column in range(4, last_day_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 4.5
    sheet.column_dimensions[last_column_letter].width = 10

    if profiles:
        sheet.auto_filter.ref = f"A2:{last_column_letter}{total_row - 1}"
    sheet.print_title_rows = "1:2"
    sheet.print_area = f"A1:{last_column_letter}{total_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.3
    sheet.page_margins.bottom = 0.3
    sheet.oddFooter.center.text = f"民國 {_roc_year(start.year)} 年 {start.month} 月｜第 &P 頁／共 &N 頁"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def csv_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def shift_detail_csv(start: date, end: date) -> bytes:
    shifts = _month_shifts(start, end)
    rows = [
        [
            shift.staff.student_number,
            shift.staff.name,
            shift.shift_date.isoformat(),
            WEEKDAYS[(shift.shift_date.weekday() + 1) % 7],
            shift.shift_type.work_location.name,
            shift.shift_type.name,
            shift.shift_type.start_time.strftime("%H:%M"),
            shift.shift_type.end_time.strftime("%H:%M"),
            f"{Decimal(str(shift.shift_type.default_hours)):g}",
            shift.status.value,
        ]
        for shift in shifts
    ]
    return csv_bytes(
        ["學號", "姓名", "日期", "星期", "地點", "班別", "開始時間", "結束時間", "時數", "狀態"],
        rows,
    )


def payroll_cost_csv(start: date, end: date) -> bytes:
    setting = get_payroll_setting(start)
    if setting is None:
        raise ValueError("此月份尚未設定薪資與保險費率。")
    hours_rows = db.session.execute(
        db.select(Shift.staff_id, db.func.sum(ShiftType.default_hours))
        .select_from(Shift)
        .join(ShiftType)
        .where(
            Shift.status == ShiftStatus.SCHEDULED,
            Shift.publication_status == ShiftPublicationStatus.PUBLISHED,
            Shift.shift_date >= start,
            Shift.shift_date < end,
        )
        .group_by(Shift.staff_id)
    ).all()
    hours_by_staff = {staff_id: Decimal(str(hours or 0)) for staff_id, hours in hours_rows}
    rows = []
    for profile in _profiles_for_month(start, end):
        item = calculate_staff_cost(
            profile=profile,
            hours=hours_by_staff.get(profile.id, Decimal("0")),
            setting=setting,
        )
        rows.append(
            [
                profile.student_number,
                profile.name,
                item["hours"],
                item["hourly_wage"],
                item["gross_wage"],
                item["labor_insurance"] + item["employment_insurance"] + item["occupational_accident"],
                item["health_insurance"],
                item["labor_pension"],
                item["employer_benefits"],
                item["employer_total"],
            ]
        )
    return csv_bytes(
        ["學號", "姓名", "時數", "時薪", "應發工資", "勞保／就保／災保", "健保", "勞退", "雇主福利成本", "雇主總成本"],
        rows,
    )


def workflow_history_csv(start: date, end: date) -> bytes:
    rows: list[list[object]] = []
    leaves = db.session.scalars(
        db.select(LeaveRequest)
        .options(
            joinedload(LeaveRequest.staff),
            joinedload(LeaveRequest.shift).joinedload(Shift.shift_type),
        )
        .where(LeaveRequest.shift.has(Shift.shift_date >= start), LeaveRequest.shift.has(Shift.shift_date < end))
        .order_by(LeaveRequest.created_at)
    ).all()
    for item in leaves:
        rows.append(
            [
                "請假",
                item.staff.student_number,
                item.staff.name,
                item.shift.shift_date.isoformat(),
                item.shift.shift_type.name,
                "",
                item.status.value,
                item.reason,
                item.note or "",
                item.created_at.isoformat(),
            ]
        )
    swaps = db.session.scalars(
        db.select(SwapRequest)
        .options(
            joinedload(SwapRequest.requester),
            joinedload(SwapRequest.target_staff),
            joinedload(SwapRequest.requester_shift).joinedload(Shift.shift_type),
        )
        .where(
            db.or_(
                SwapRequest.requester_shift.has(
                    db.and_(Shift.shift_date >= start, Shift.shift_date < end)
                ),
                SwapRequest.target_shift.has(
                    db.and_(Shift.shift_date >= start, Shift.shift_date < end)
                ),
            )
        )
        .order_by(SwapRequest.created_at)
    ).all()
    for item in swaps:
        rows.append(
            [
                "換班",
                item.requester.student_number,
                item.requester.name,
                item.requester_shift.shift_date.isoformat(),
                item.requester_shift.shift_type.name,
                f"{item.target_staff.student_number} {item.target_staff.name}",
                item.display_status,
                "",
                item.note or "",
                item.created_at.isoformat(),
            ]
        )
    rows.sort(key=lambda row: row[-1])
    return csv_bytes(
        ["類型", "申請人學號", "申請人", "排班日期", "班別", "對象", "狀態", "原因", "備註", "申請時間"],
        rows,
    )


def document_expiry_csv() -> bytes:
    confirmed = db.session.scalars(
        db.select(StaffDocument)
        .where(StaffDocument.status == DocumentStatus.CONFIRMED)
        .order_by(StaffDocument.uploaded_at.desc())
    ).all()
    page_sets: dict[tuple[int, str], set[DocumentPageKind]] = defaultdict(set)
    for document in confirmed:
        page_sets[(document.staff_id, document.document_type.value)].add(document.page_kind)
    rows = []
    for profile in _active_profiles():
        residence_pages = page_sets.get((profile.id, "RESIDENCE_PERMIT"), set())
        work_pages = page_sets.get((profile.id, "WORK_PERMIT"), set())
        rows.append(
            [
                profile.student_number,
                profile.name,
                profile.nationality,
                profile.residence_expiry.isoformat() if profile.residence_expiry else "",
                expiry_state(profile.residence_expiry)["label"],
                "完整" if {DocumentPageKind.RESIDENCE_FRONT, DocumentPageKind.RESIDENCE_BACK}.issubset(residence_pages) else "不完整",
                profile.work_permit_start.isoformat() if profile.work_permit_start else "",
                profile.work_permit_expiry.isoformat() if profile.work_permit_expiry else "",
                expiry_state(profile.work_permit_expiry)["label"],
                f"{len(work_pages)} 頁" if work_pages else "未上傳",
            ]
        )
    return csv_bytes(
        ["學號", "姓名", "國籍", "居留證截止日", "居留證狀態", "居留證正反面", "工作證開始日", "工作證截止日", "工作證狀態", "工作證頁數"],
        rows,
    )
