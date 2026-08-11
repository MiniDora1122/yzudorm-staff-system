from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app.extensions import db
from app.models import AuditLog, Shift, ShiftStatus, ShiftType, StaffProfile, User

from .conftest import login


def add_august_shift(app):
    with app.app_context():
        profile = db.session.scalar(
            db.select(StaffProfile).where(StaffProfile.student_number == "TEST001")
        )
        shift_type = db.session.scalar(db.select(ShiftType).where(ShiftType.code == "TEST_AM"))
        admin = db.session.scalar(db.select(User).where(User.username == "admin-test"))
        db.session.add(
            Shift(
                shift_date=date(2026, 8, 10),
                shift_type_id=shift_type.id,
                staff_id=profile.id,
                status=ShiftStatus.SCHEDULED,
                created_by=admin.id,
            )
        )
        db.session.commit()


def test_reference_style_monthly_report_contains_student_number_hours_and_formulas(client, app):
    add_august_shift(app)
    login(client)
    response = client.get("/admin/reports/monthly-hours.xlsx?month=2026-08")
    assert response.status_code == 200
    assert "spreadsheetml" in response.mimetype
    workbook = load_workbook(BytesIO(response.data), data_only=False)
    sheet = workbook.active
    values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert any(isinstance(value, str) and "TEST001" in value for value in values)
    assert 10 in values
    assert 4 in values
    assert any(isinstance(value, str) and value.startswith("=SUM(") for value in values)
    assert sheet.sheet_view.showGridLines is False
    assert sheet.page_setup.fitToWidth == 1
    with app.app_context():
        assert db.session.scalar(
            db.select(db.func.count()).select_from(AuditLog).where(
                AuditLog.action == "REPORT_DOWNLOADED"
            )
        ) == 1


def test_reports_are_admin_only_and_csv_has_utf8_bom_and_student_number(client, app):
    add_august_shift(app)
    login(client, "student-test", "StudentTest!2026")
    assert client.get("/admin/reports").status_code == 403
    client.post("/auth/logout")
    login(client)
    response = client.get("/admin/reports/shifts.csv?month=2026-08")
    assert response.status_code == 200
    assert response.data.startswith(b"\xef\xbb\xbf")
    decoded = response.data.decode("utf-8-sig")
    assert "學號" in decoded
    assert "TEST001" in decoded
    assert "辦公室" in decoded


def test_admin_can_update_student_number_but_duplicate_is_rejected(client, app):
    login(client)
    with app.app_context():
        profile = db.session.scalar(
            db.select(StaffProfile).where(StaffProfile.student_number == "TEST001")
        )
        profile_id = profile.id
    response = client.post(
        f"/admin/staff/{profile_id}",
        data={
            "name": "測試學生",
            "student_number": "NEW001",
            "email": "student@example.edu.tw",
            "phone": "0912345678",
            "nationality": "台灣",
        },
        follow_redirects=True,
    )
    assert "學號已更新".encode() in response.data
    with app.app_context():
        assert db.session.get(StaffProfile, profile_id).student_number == "NEW001"

    duplicate = client.post(
        f"/admin/staff/{profile_id}",
        data={
            "name": "測試學生",
            "student_number": "TEST002",
            "email": "student@example.edu.tw",
            "phone": "0912345678",
            "nationality": "台灣",
        },
        follow_redirects=True,
    )
    assert "此學號已由其他工讀生使用".encode() in duplicate.data
    with app.app_context():
        assert db.session.get(StaffProfile, profile_id).student_number == "NEW001"


def test_ocr_route_is_removed(client):
    login(client, "student-test", "StudentTest!2026")
    assert client.post("/student/documents/1/ocr").status_code == 404
